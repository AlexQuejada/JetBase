"""
Manejador de lectura robusta de archivos Excel.
"""
import io
import warnings
import pandas as pd
from typing import Optional, List

from .common import clean_dataframe, FileReadingError


# Import opcional de openpyxl para manejo avanzado de celdas mergeadas
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def _handle_merged_cells(workbook, sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    """
    Procesa celdas mergeadas en Excel, rellenando valores hacia abajo.

    Args:
        workbook: Objeto workbook de openpyxl
        sheet_name: Nombre de la hoja
        df: DataFrame a procesar

    Returns:
        DataFrame con celdas mergeadas expandidas
    """
    if not OPENPYXL_AVAILABLE:
        return df

    try:
        if hasattr(workbook, 'sheetnames'):
            ws = workbook[sheet_name]
            if ws.merged_cells.ranges:
                for merged_range in ws.merged_cells.ranges:
                    min_row = merged_range.min_row - 1
                    max_row = merged_range.max_row - 1
                    min_col = merged_range.min_col - 1
                    max_col = merged_range.max_col - 1

                    top_left_value = ws.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    ).value

                    if min_row < len(df) and min_col < len(df.columns):
                        for row in range(min_row, min(max_row + 1, len(df))):
                            for col in range(min_col, min(max_col + 1, len(df.columns))):
                                if pd.isna(df.iloc[row, col]) or df.iloc[row, col] == '':
                                    df.iloc[row, col] = top_left_value
    except Exception:
        pass

    return df


async def read_excel_robust(
    contents: bytes,
    sheet_name: Optional[str] = None,
    combine_sheets: bool = False
) -> Optional[pd.DataFrame]:
    """
    Lee un archivo Excel (.xlsx, .xls, .xlsm) de forma robusta.

    Maneja:
    - Múltiples formatos (.xlsx moderno, .xls antiguo)
    - Selección de hoja específica o primera con datos
    - Combinación de múltiples hojas
    - Celdas mergeadas (expandidas automáticamente)

    Args:
        contents: Contenido binario del archivo
        sheet_name: Nombre de hoja específica (None = auto)
        combine_sheets: Si True, combina todas las hojas

    Returns:
        DataFrame o None si falla
    """
    if not contents or len(contents) < 10:
        return None

    # Determinar motor según disponibilidad
    engines_to_try = ['openpyxl', 'xlrd', None]
    xl_file = None

    for engine in engines_to_try:
        try:
            xl_file = pd.ExcelFile(io.BytesIO(contents), engine=engine)
            break
        except Exception:
            continue

    if xl_file is None:
        return None

    try:
        sheet_names = xl_file.sheet_names
        if not sheet_names:
            return None

        dfs = []
        sheets_to_process = []

        if combine_sheets:
            sheets_to_process = sheet_names
        elif sheet_name:
            if sheet_name in sheet_names:
                sheets_to_process = [sheet_name]
            else:
                # Búsqueda case-insensitive
                sheet_lower = sheet_name.lower()
                for sn in sheet_names:
                    if sn.lower() == sheet_lower:
                        sheets_to_process = [sn]
                        break
                if not sheets_to_process:
                    return None
        else:
            # Encontrar primera hoja con datos
            for sn in sheet_names:
                try:
                    df = pd.read_excel(xl_file, sheet_name=sn)
                    if not df.empty and len(df.columns) > 0:
                        sheets_to_process = [sn]
                        break
                except Exception:
                    continue
            if not sheets_to_process:
                sheets_to_process = [sheet_names[0]]

        # Leer cada hoja
        for sn in sheets_to_process:
            try:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    df = pd.read_excel(xl_file, sheet_name=sn)

                if df is not None and not df.empty:
                    df = clean_dataframe(df)

                    # Manejar celdas mergeadas
                    if OPENPYXL_AVAILABLE:
                        try:
                            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
                            df = _handle_merged_cells(wb, sn, df)
                        except Exception:
                            pass

                    # Agregar origen si combina múltiples hojas
                    if combine_sheets and len(sheets_to_process) > 1:
                        df['__sheet_name__'] = sn

                    dfs.append(df)
            except Exception:
                continue

        if not dfs:
            return None

        if len(dfs) == 1:
            return dfs[0]

        # Combinar múltiples hojas
        all_columns = set()
        for df in dfs:
            all_columns.update(df.columns)

        aligned_dfs = []
        for df in dfs:
            for col in all_columns:
                if col not in df.columns:
                    df[col] = None
            aligned_dfs.append(df[list(all_columns)])

        return pd.concat(aligned_dfs, ignore_index=True)

    except Exception:
        return None
